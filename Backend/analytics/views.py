# analytics/views.py
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.utils import timezone
from datetime import timedelta
from django.db.models import Count, Q, Sum, Avg, F
from projects.models import Project
from tasks.models import Task
from accounts.models import User
from accounts.permissions import IsAdminOrManager
import logging
import csv
import io
from django.http import FileResponse, HttpResponse
import json

logger = logging.getLogger(__name__)

# Try to import reportlab and openpyxl (optional dependencies)
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    logger.warning("reportlab not installed. PDF export disabled.")

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not installed. Excel export disabled.")


class AnalyticsView(APIView):
    """Analytics data for Admin - all projects."""
    permission_classes = [IsAuthenticated, IsAdminOrManager]

    def get(self, request):
        try:
            now = timezone.now().date()

            # ✅ Check if user is Admin or Manager
            is_admin = request.user.role == "Admin"
            is_manager = request.user.role == "Manager"

            # ✅ Get filter parameters
            status_filter = request.query_params.get('status', None)
            priority_filter = request.query_params.get('priority', None)
            project_filter = request.query_params.get('project', None)

            # ✅ Filter based on role
            if is_manager:
                projects = Project.objects.filter(manager=request.user)
                project_ids = projects.values_list('id', flat=True)
                tasks = Task.objects.filter(project__in=project_ids)
                users = User.objects.filter(tasks__project__in=project_ids).distinct()
            else:
                projects = Project.objects.all()
                tasks = Task.objects.all()
                users = User.objects.all()

            # ✅ Apply status filter to projects
            if status_filter and status_filter != 'all':
                if status_filter == 'active':
                    projects = projects.filter(status='active')
                elif status_filter == 'completed':
                    projects = projects.filter(status='completed')
                elif status_filter == 'pending':
                    projects = projects.filter(status='pending')
                elif status_filter == 'archived':
                    projects = projects.filter(status='archived')
                
            # ✅ Apply priority filter to tasks
            if priority_filter and priority_filter != 'all':
                tasks = tasks.filter(priority=priority_filter)
                
            # ✅ Fix: Handle project filter properly - check if it's a valid ID
            if project_filter and project_filter != 'all':
                try:
                    # Try to convert to integer (for project ID)
                    project_id = int(project_filter)
                    tasks = tasks.filter(project_id=project_id)
                    projects = projects.filter(id=project_id)
                except ValueError:
                    # If it's a string like 'project-1', try to find by name or skip
                    logger.warning(f"Invalid project filter: {project_filter}")
                    # Try to find project by name (optional)
                    project_obj = projects.filter(name__icontains=project_filter).first()
                    if project_obj:
                        tasks = tasks.filter(project_id=project_obj.id)
                        projects = projects.filter(id=project_obj.id)

            # ==================== 📊 OVERVIEW STATS ====================
            
            total_projects = projects.count()
            active_projects = projects.filter(status="active").count()
            completed_projects = projects.filter(status="completed").count()
            archived_projects = projects.filter(status="archived").count()
            
            total_tasks = tasks.count()
            completed_tasks = tasks.filter(status="completed").count()
            in_progress_tasks = tasks.filter(status="in_progress").count()
            pending_tasks = tasks.filter(status="pending").count()
            
            overdue_tasks = tasks.filter(
                due_date__lt=now,
                status__in=["pending", "in_progress"]
            ).count()
            
            completion_rate = (completed_tasks / total_tasks * 100) if total_tasks > 0 else 0

            team_members = users.count()

            # ==================== 📊 MONTHLY STATS ====================
            monthly_stats = []
            for i in range(6, -1, -1):
                month_date = now - timedelta(days=30 * i)
                month_start = month_date.replace(day=1)
                if i > 0:
                    month_end = (month_date.replace(day=1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)
                else:
                    month_end = now
                
                month_projects = projects.filter(
                    start_date__gte=month_start,
                    start_date__lte=month_end
                )
                
                month_tasks = tasks.filter(
                    due_date__gte=month_start,
                    due_date__lte=month_end
                )
                month_completed = tasks.filter(
                    status="completed",
                    due_date__gte=month_start,
                    due_date__lte=month_end
                )
                
                monthly_stats.append({
                    "month": month_start.strftime("%b"),
                    "projects": month_projects.count(),
                    "tasks": month_tasks.count(),
                    "completed": month_completed.count(),
                })

            # ==================== 📊 PROJECT STATUS ====================
            
            project_status = [
                {"name": "Active", "value": active_projects, "color": "#3B82F6"},
                {"name": "Completed", "value": completed_projects, "color": "#10B981"},
                {"name": "Archived", "value": archived_projects, "color": "#6B7280"},
            ]

            # ==================== 📊 TASK DISTRIBUTION ====================
            
            task_distribution = [
                {"name": "Pending", "value": pending_tasks, "color": "#F59E0B"},
                {"name": "In Progress", "value": in_progress_tasks, "color": "#3B82F6"},
                {"name": "Completed", "value": completed_tasks, "color": "#10B981"},
            ]

            # ==================== 📊 WEEKLY ACTIVITY ====================
            weekly_activity = []
            days = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
            for i, day in enumerate(days):
                day_date = now - timedelta(days=(now.weekday() - i) % 7)
                day_tasks = tasks.filter(
                    due_date=day_date
                )
                day_completed = tasks.filter(
                    status="completed",
                    due_date=day_date
                )
                weekly_activity.append({
                    "day": day,
                    "tasks": day_tasks.count(),
                    "completed": day_completed.count(),
                })

            # ==================== 📊 PERFORMANCE METRICS ====================
            
            performance = {
                "efficiency": min(100, int((completed_tasks / (total_tasks + 1)) * 100)),
                "productivity": min(100, int((total_tasks / (team_members + 1)) * 20) if team_members > 0 else 0),
                "quality": min(100, int((completed_tasks / (total_tasks + 1)) * 95)),
                "timeliness": min(100, int(((total_tasks - overdue_tasks) / (total_tasks + 1)) * 100)),
            }

            # ==================== 📊 RECENT ACTIVITY ====================
            
            recent_activity = []
            recent_tasks = tasks.order_by('-id')[:10]
            for task in recent_tasks:
                recent_activity.append({
                    "id": task.id,
                    "user": task.assigned_to.username if task.assigned_to else "Unassigned",
                    "action": "completed" if task.status == "completed" else "updated" if task.status == "in_progress" else "created",
                    "task": task.title,
                    "type": "complete" if task.status == "completed" else "update" if task.status == "in_progress" else "create",
                    "timestamp": timezone.now().isoformat(),
                })

            # ==================== 📦 BUILD RESPONSE ====================
            
            response_data = {
                "overview": {
                    "totalProjects": total_projects,
                    "activeProjects": active_projects,
                    "completedProjects": completed_projects,
                    "archivedProjects": archived_projects,
                    "totalTasks": total_tasks,
                    "completedTasks": completed_tasks,
                    "inProgressTasks": in_progress_tasks,
                    "pendingTasks": pending_tasks,
                    "overdueTasks": overdue_tasks,
                    "teamMembers": team_members,
                    "completionRate": round(completion_rate, 1),
                    "avgTaskDuration": "N/A",
                },
                "monthlyStats": monthly_stats,
                "projectStatus": project_status,
                "taskDistribution": task_distribution,
                "weeklyActivity": weekly_activity,
                "performance": performance,
                "recentActivity": recent_activity,
            }

            return Response(response_data)

        except Exception as e:
            logger.error(f"Analytics error: {e}")
            import traceback
            traceback.print_exc()
            return Response({
                "error": "Failed to generate analytics",
                "detail": str(e)
            }, status=500)


class ManagerAnalyticsView(APIView):
    """Analytics data for Managers - only their projects."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            projects = Project.objects.filter(manager=request.user)
            project_ids = projects.values_list('id', flat=True)
            tasks = Task.objects.filter(project__in=project_ids)
            
            now = timezone.now().date()
            
            total_projects = projects.count()
            active_projects = projects.filter(status="active").count()
            completed_projects = projects.filter(status="completed").count()
            
            total_tasks = tasks.count()
            completed_tasks = tasks.filter(status="completed").count()
            in_progress_tasks = tasks.filter(status="in_progress").count()
            pending_tasks = tasks.filter(status="pending").count()
            overdue_tasks = tasks.filter(
                due_date__lt=now,
                status__in=["pending", "in_progress"]
            ).count()
            
            completion_rate = (completed_tasks / total_tasks * 100) if total_tasks > 0 else 0
            
            team_members = User.objects.filter(tasks__project__in=project_ids).distinct().count()
            
            project_data = []
            for project in projects:
                task_count = Task.objects.filter(project=project).count()
                project_data.append({
                    "name": project.name,
                    "value": task_count,
                })
            
            weekly_trend = []
            for i in range(6, -1, -1):
                date = now - timedelta(days=i)
                day_tasks = tasks.filter(due_date=date)
                weekly_trend.append({
                    "day": date.strftime("%a"),
                    "value": day_tasks.count(),
                })
            
            return Response({
                "totalProjects": total_projects,
                "activeProjects": active_projects,
                "completedProjects": completed_projects,
                "totalTasks": total_tasks,
                "completedTasks": completed_tasks,
                "overdueTasks": overdue_tasks,
                "inProgressTasks": in_progress_tasks,
                "teamMembers": team_members,
                "completionRate": round(completion_rate, 1),
                "projectData": project_data,
                "weeklyTrend": weekly_trend,
            })
            
        except Exception as e:
            logger.error(f"Manager Analytics error: {e}")
            return Response({"error": str(e)}, status=500)


# ============================================
# 📊 EXPORT ANALYTICS
# ============================================

class AnalyticsExportView(APIView):
    """Export analytics data as PDF, CSV, or Excel."""
    permission_classes = [IsAuthenticated, IsAdminOrManager]

    def post(self, request):
        try:
            report_type = request.data.get('reportType', 'summary')
            format_type = request.data.get('format', 'pdf')
            include_charts = request.data.get('includeCharts', True)
            filters = request.data.get('filters', {})
            analytics_data = request.data.get('data', {})

            # ✅ Get fresh data from database
            fresh_data = self.get_fresh_data(request, filters)

            if format_type == 'csv':
                return self.export_csv(fresh_data, report_type)
            elif format_type == 'excel':
                return self.export_excel(fresh_data, report_type)
            else:
                return self.export_pdf(fresh_data, report_type, include_charts)

        except Exception as e:
            logger.error(f"Export error: {e}")
            import traceback
            traceback.print_exc()
            return Response({"error": str(e)}, status=500)

    def get_fresh_data(self, request, filters):
        """Get fresh data from database with filters applied."""
        now = timezone.now().date()
        
        # Get projects based on role
        if request.user.role == "Manager":
            projects = Project.objects.filter(manager=request.user)
            project_ids = projects.values_list('id', flat=True)
            tasks = Task.objects.filter(project__in=project_ids)
        else:
            projects = Project.objects.all()
            tasks = Task.objects.all()

        # Apply filters
        if filters.get('status') and filters.get('status') != 'all':
            if filters['status'] == 'active':
                projects = projects.filter(status='active')
            elif filters['status'] == 'completed':
                projects = projects.filter(status='completed')
                tasks = tasks.filter(status='completed')
            elif filters['status'] == 'pending':
                tasks = tasks.filter(status='pending')
                
        if filters.get('priority') and filters.get('priority') != 'all':
            tasks = tasks.filter(priority=filters['priority'])
            
        if filters.get('project') and filters.get('project') != 'all':
            try:
                project_id = int(filters['project'])
                tasks = tasks.filter(project_id=project_id)
                projects = projects.filter(id=project_id)
            except ValueError:
                pass

        # Build fresh data
        return {
            "overview": {
                "totalProjects": projects.count(),
                "activeProjects": projects.filter(status="active").count(),
                "completedProjects": projects.filter(status="completed").count(),
                "archivedProjects": projects.filter(status="archived").count(),
                "totalTasks": tasks.count(),
                "completedTasks": tasks.filter(status="completed").count(),
                "inProgressTasks": tasks.filter(status="in_progress").count(),
                "pendingTasks": tasks.filter(status="pending").count(),
                "overdueTasks": tasks.filter(
                    due_date__lt=now,
                    status__in=["pending", "in_progress"]
                ).count(),
                "completionRate": round((tasks.filter(status="completed").count() / max(tasks.count(), 1)) * 100, 1),
            },
            "taskDistribution": [
                {"name": "Pending", "value": tasks.filter(status="pending").count()},
                {"name": "In Progress", "value": tasks.filter(status="in_progress").count()},
                {"name": "Completed", "value": tasks.filter(status="completed").count()},
            ],
            "projectStatus": [
                {"name": "Active", "value": projects.filter(status="active").count()},
                {"name": "Completed", "value": projects.filter(status="completed").count()},
                {"name": "Archived", "value": projects.filter(status="archived").count()},
            ],
            "monthlyStats": self.get_monthly_stats(projects, tasks, now),
            "weeklyActivity": self.get_weekly_activity(tasks, now),
        }

    def get_monthly_stats(self, projects, tasks, now):
        """Generate monthly statistics."""
        monthly_stats = []
        for i in range(6, -1, -1):
            month_date = now - timedelta(days=30 * i)
            month_start = month_date.replace(day=1)
            if i > 0:
                month_end = (month_date.replace(day=1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)
            else:
                month_end = now
            
            month_projects = projects.filter(
                start_date__gte=month_start,
                start_date__lte=month_end
            )
            
            month_tasks = tasks.filter(
                due_date__gte=month_start,
                due_date__lte=month_end
            )
            month_completed = tasks.filter(
                status="completed",
                due_date__gte=month_start,
                due_date__lte=month_end
            )
            
            monthly_stats.append({
                "month": month_start.strftime("%b"),
                "projects": month_projects.count(),
                "tasks": month_tasks.count(),
                "completed": month_completed.count(),
            })
        return monthly_stats

    def get_weekly_activity(self, tasks, now):
        """Generate weekly activity."""
        weekly_activity = []
        days = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
        for i, day in enumerate(days):
            day_date = now - timedelta(days=(now.weekday() - i) % 7)
            day_tasks = tasks.filter(due_date=day_date)
            day_completed = tasks.filter(status="completed", due_date=day_date)
            weekly_activity.append({
                "day": day,
                "tasks": day_tasks.count(),
                "completed": day_completed.count(),
            })
        return weekly_activity

    def export_csv(self, data, report_type):
        """Export as CSV"""
        output = io.StringIO()
        writer = csv.writer(output)

        # Write header
        if report_type == 'summary':
            writer.writerow(['SmartTask Analytics Report'])
            writer.writerow(['Generated:', timezone.now().strftime('%Y-%m-%d %H:%M')])
            writer.writerow([])
            writer.writerow(['Metric', 'Value'])
            overview = data.get('overview', {})
            for key, value in overview.items():
                writer.writerow([key.replace('_', ' ').title(), value])

        elif report_type == 'tasks':
            writer.writerow(['Task Report'])
            writer.writerow(['Generated:', timezone.now().strftime('%Y-%m-%d %H:%M')])
            writer.writerow([])
            writer.writerow(['Status', 'Count'])
            task_dist = data.get('taskDistribution', [])
            for item in task_dist:
                writer.writerow([item['name'], item['value']])

        elif report_type == 'projects':
            writer.writerow(['Project Report'])
            writer.writerow(['Generated:', timezone.now().strftime('%Y-%m-%d %H:%M')])
            writer.writerow([])
            writer.writerow(['Status', 'Count'])
            project_status = data.get('projectStatus', [])
            for item in project_status:
                writer.writerow([item['name'], item['value']])

        else:  # detailed
            writer.writerow(['Detailed Analytics Report'])
            writer.writerow(['Generated:', timezone.now().strftime('%Y-%m-%d %H:%M')])
            writer.writerow([])
            writer.writerow(['Metric', 'Value'])
            overview = data.get('overview', {})
            for key, value in overview.items():
                writer.writerow([key.replace('_', ' ').title(), value])

        response = HttpResponse(output.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="analytics_report_{timezone.now().strftime("%Y%m%d_%H%M")}.csv"'
        return response

    def export_excel(self, data, report_type):
        """Export as Excel (using openpyxl)"""
        if not OPENPYXL_AVAILABLE:
            return Response({"error": "openpyxl not installed. Please install it: pip install openpyxl"}, status=500)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Analytics Report"

        # Title
        title_cell = ws['A1']
        title_cell.value = f"SmartTask Analytics - {report_type.title()} Report"
        title_cell.font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')

        # Date
        ws['A2'] = f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')}"
        ws.merge_cells('A2:D2')

        # Headers
        row = 4
        if report_type == 'summary' or report_type == 'detailed':
            headers = ['Metric', 'Value']
            ws.append(headers)
            overview = data.get('overview', {})
            for key, value in overview.items():
                ws.append([key.replace('_', ' ').title(), value])

        elif report_type == 'tasks':
            headers = ['Status', 'Count']
            ws.append(headers)
            task_dist = data.get('taskDistribution', [])
            for item in task_dist:
                ws.append([item['name'], item['value']])

        elif report_type == 'projects':
            headers = ['Status', 'Count']
            ws.append(headers)
            project_status = data.get('projectStatus', [])
            for item in project_status:
                ws.append([item['name'], item['value']])

        # Style headers
        for col in range(1, 3):
            cell = ws.cell(row=4, column=col)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')

        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="analytics_report_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
        wb.save(response)
        return response

    def export_pdf(self, data, report_type, include_charts):
        """Export as PDF"""
        if not REPORTLAB_AVAILABLE:
            return Response({"error": "reportlab not installed. Please install it: pip install reportlab"}, status=500)

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=40, bottomMargin=40)
        styles = getSampleStyleSheet()
        elements = []

        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.HexColor('#4F46E5'),
            alignment=TA_CENTER,
            spaceAfter=30
        )
        elements.append(Paragraph(f"SmartTask Analytics Report", title_style))
        elements.append(Paragraph(f"<b>Report Type:</b> {report_type.title()}", styles['Normal']))
        elements.append(Paragraph(f"<b>Generated:</b> {timezone.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
        elements.append(Spacer(1, 20))

        # Overview data
        overview = data.get('overview', {})
        if overview:
            elements.append(Paragraph("<b>Overview</b>", styles['Heading2']))
            elements.append(Spacer(1, 10))
            
            table_data = [['Metric', 'Value']]
            for key, value in overview.items():
                table_data.append([key.replace('_', ' ').title(), str(value)])
            
            table = Table(table_data, colWidths=[200, 100])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(table)
            elements.append(Spacer(1, 20))

        # Monthly Stats
        monthly_stats = data.get('monthlyStats', [])
        if monthly_stats:
            elements.append(Paragraph("<b>Monthly Activity</b>", styles['Heading2']))
            elements.append(Spacer(1, 10))
            
            table_data = [['Month', 'Projects', 'Tasks', 'Completed']]
            for item in monthly_stats:
                table_data.append([
                    item['month'],
                    str(item['projects']),
                    str(item['tasks']),
                    str(item['completed'])
                ])
            
            table = Table(table_data, colWidths=[80, 80, 80, 80])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3B82F6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(table)
            elements.append(Spacer(1, 20))

        # Task distribution
        task_dist = data.get('taskDistribution', [])
        if task_dist:
            elements.append(Paragraph("<b>Task Distribution</b>", styles['Heading2']))
            elements.append(Spacer(1, 10))
            
            table_data = [['Status', 'Count']]
            for item in task_dist:
                table_data.append([item['name'], str(item['value'])])
            
            table = Table(table_data, colWidths=[150, 100])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10B981')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(table)

        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        
        response = FileResponse(buffer, as_attachment=True, 
                               filename=f'analytics_report_{timezone.now().strftime("%Y%m%d_%H%M")}.pdf')
        return response